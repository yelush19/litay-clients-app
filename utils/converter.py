import io
import re
from datetime import date
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

VAT_RATE = 0.18

# חשבון זכות לפי סוג עמודה
KNOWN_COLUMNS = {
    "שכט":     (5000, False),
    "נוטריון": (5004, False),
    "הוצאות":  (5002, True),
}

INVOICE_HEADERS = [
    "תאריך", "חשבון חובה 1", "חשבון זכות 1", "חשבון זכות 2",
    "פרטים", "אסמכתא", "לקוח-סכום חובה 1", "הכנסה-סכום זכות 1", "מעמע-סכום זכות 2",
]
RECEIPT_HEADERS = [
    "תאריך", "חשבון חובה 1", "חשבון זכות",
    "פרטים", "אסמכתא", "בנק-סכום חובה 1", "לקוח-סכום זכות 2",
]
ALLOCATION_HEADERS = [
    "מספר חשבונית", "תאריך", "שם לקוח", "ת.ז / מס.ע.מ",
    "סכום לפני מע\"מ", "מספר הקצאה",
]


def detect_month_label(filename: str, df: pd.DataFrame) -> str:
    match = re.search(r'[_\-\s](\d{1,2})[_\-\s\.](\d{4})', filename)
    if match:
        return f"{int(match.group(1)):02d}.{match.group(2)}"
    for i in range(1, min(5, len(df))):
        val = df.iloc[i, 1]
        if pd.notna(val) and hasattr(val, "month"):
            return f"{val.month:02d}.{val.year}"
    return "חדש"


def _clean(val) -> float:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return 0.0
    try:
        return float(val)
    except Exception:
        return 0.0


def _allocation_threshold(txn_date) -> float:
    """סף הקצאה לפי תאריך"""
    cutoff = date(2026, 6, 1)
    d = txn_date if isinstance(txn_date, date) else txn_date.date()
    return 5000.0 if d >= cutoff else 10000.0


def convert_income_file(df: pd.DataFrame, clients_dict: dict, extra_columns: dict,
                        clients_id: dict = None):
    """
    Returns (invoice_rows, receipt_rows, allocation_rows, unmatched, unknown_cols)
    clients_id: {name: id_number}
    """
    from utils.matcher import flexible_match

    invoice_rows    = []
    receipt_rows    = []
    allocation_rows = []
    unmatched       = []
    unknown_columns = set()

    for i in range(1, len(df)):
        row = df.iloc[i]

        total       = _clean(row.iloc[0])
        txn_date    = row.iloc[1]
        vat_file    = _clean(row.iloc[2])   # מעמ מהקובץ
        notary      = _clean(row.iloc[4]) if len(row) > 4 else 0
        expenses    = _clean(row.iloc[5]) if len(row) > 5 else 0
        fee         = _clean(row.iloc[6]) if len(row) > 6 else 0
        misc        = _clean(row.iloc[7]) if len(row) > 7 else 0
        client_name = str(row.iloc[9]).strip() if len(row) > 9 else ""
        import re as _re
        _inv_raw = row.iloc[10] if len(row) > 10 else None
        invoice_num = _re.sub(r'[^0-9]', '', str(_inv_raw or ''))[:9] if _inv_raw is not None else ''


        if pd.isna(txn_date) or not hasattr(txn_date, "month"):
            continue
        if not client_name or client_name in ("nan", "לקוח", ""):
            continue

        txn_date_obj = txn_date.date() if hasattr(txn_date, "date") else txn_date
        txn_date_str = f"{txn_date_obj.day:02d}/{txn_date_obj.month:02d}/{txn_date_obj.year}"

        account, ratio, matched_name = flexible_match(client_name, clients_dict)
        if account is None:
            if client_name not in unmatched:
                unmatched.append(client_name)
            continue

        col_data = {}
        if fee      > 0: col_data["שכט"]     = fee
        if notary   > 0: col_data["נוטריון"] = notary
        if expenses > 0: col_data["הוצאות"]  = expenses
        if misc     > 0: col_data["שונות"]   = misc

        for col_name in col_data:
            if col_name not in KNOWN_COLUMNS and col_name not in extra_columns:
                unknown_columns.add(col_name)

        if any(c not in KNOWN_COLUMNS and c not in extra_columns for c in col_data):
            continue

        # חישוב מע"מ — שימוש בסכום מהקובץ (מדויק יותר)
        taxable_cols = {k: v for k, v in col_data.items()
                        if not (KNOWN_COLUMNS.get(k, (None, None))[1] or
                                extra_columns.get(k, {}).get("vat_exempt", False))}
        total_taxable = sum(taxable_cols.values())
        accum_vat = 0.0

        taxable_list = list(taxable_cols.items())
        for j, (col_name, col_value) in enumerate(taxable_list):
            is_last = (j == len(taxable_list) - 1)
            if total_taxable > 0:
                if is_last:
                    col_vat = round(vat_file - accum_vat, 2)
                else:
                    col_vat = round(vat_file * col_value / total_taxable, 2)
                    accum_vat += col_vat
            else:
                col_vat = 0.0

            credit_account = KNOWN_COLUMNS.get(col_name, (extra_columns.get(col_name, {}).get("account"), None))[0]
            row_income = round(col_value, 2)
            row_total  = round(col_value + col_vat, 2)

            # ── בדיקת איזון: חובה = זכות1 + זכות2 ──
            diff = round(row_total - row_income - col_vat, 2)
            if abs(diff) > 0:
                row_income = round(row_income + diff, 2)

            invoice_rows.append([
                txn_date_str, account, credit_account, 9001,
                client_name, invoice_num,
                row_total, row_income, col_vat,
            ])

        # עמודות פטורות ממע"מ
        for col_name in col_data:
            cfg = KNOWN_COLUMNS.get(col_name) or (extra_columns.get(col_name, {}).get("account"),
                                                   extra_columns.get(col_name, {}).get("vat_exempt", False))
            credit_account, is_exempt = cfg if isinstance(cfg, tuple) else (cfg["account"], cfg["vat_exempt"])
            if not is_exempt:
                continue
            col_value = col_data[col_name]
            invoice_rows.append([
                txn_date_str, account, credit_account, 9001,
                client_name, invoice_num,
                round(col_value, 2), round(col_value, 2), 0.0,
            ])

        # ── קבלה ──
        if total > 0:
            receipt_rows.append([
                txn_date_str, 1200, account,
                client_name, invoice_num,
                total, total,
            ])

        # ── דוח הקצאה ──
        net_taxable = fee + notary
        threshold   = _allocation_threshold(txn_date_obj)
        if net_taxable >= threshold:
            id_number = (clients_id or {}).get(client_name, "")
            if not id_number and matched_name:
                id_number = (clients_id or {}).get(matched_name, "")
            allocation_rows.append([
                invoice_num, txn_date_str, client_name,
                id_number, net_taxable, "",
            ])

    return invoice_rows, receipt_rows, allocation_rows, unmatched, list(unknown_columns)


# ── פורמט Excel ────────────────────────────────────────────

HDR_FILL = PatternFill("solid", fgColor="1E4D2B")
HDR_FONT = Font(bold=True, color="FFFFFF")

# רוחבי עמודות קבועים
INV_WIDTHS = [13, 15, 15, 15, 26, 10, 18, 18, 18]
REC_WIDTHS = [13, 15, 15, 26, 10, 18, 18]
ALC_WIDTHS = [12, 13, 26, 15, 18, 18]

# עמודות שצריכות פורמט TEXT (אינדקס 1-based)
INV_TEXT_COLS = {2, 3, 4, 6}   # חשבון חובה 1, זכות 1, זכות 2, אסמכתא
REC_TEXT_COLS = {2, 3, 5}      # חשבון חובה 1, זכות, אסמכתא


def _format_sheet(ws, headers, rows, col_widths, text_cols: set):
    # כותרת
    ws.append(headers)
    for i, cell in enumerate(ws[1], 1):
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
    ws.row_dimensions[1].height = 22

    # נתונים
    for row in rows:
        ws.append(row)

    # עיצוב תאים
    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)

            if c_idx == 1:                          # תאריך — TEXT
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center")

            elif c_idx in text_cols:               # חשבון / אסמכתא — TEXT
                if cell.value is not None:
                    cell.value = str(int(cell.value)) if isinstance(cell.value, float) else str(cell.value)
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center")

            elif isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="left")

            else:
                cell.alignment = Alignment(horizontal="right")


def _add_named_range(wb, ws, headers, num_rows, range_name):
    """הגדרת שם לטווח הנתונים"""
    if num_rows == 0:
        return
    last_col = get_column_letter(len(headers))
    last_row = num_rows + 1
    ref = f"'{ws.title}'!$A$1:${last_col}${last_row}"
    wb.defined_names[range_name] = DefinedName(range_name, attr_text=ref)


def create_excel_output(invoice_rows, receipt_rows, month_label: str) -> io.BytesIO:
    wb = Workbook()

    ws_inv = wb.active
    ws_inv.title = month_label
    _format_sheet(ws_inv, INVOICE_HEADERS, invoice_rows, INV_WIDTHS, INV_TEXT_COLS)
    _add_named_range(wb, ws_inv, INVOICE_HEADERS, len(invoice_rows), "Database")

    ws_rec = wb.create_sheet(f"{month_label} קבלות")
    _format_sheet(ws_rec, RECEIPT_HEADERS, receipt_rows, REC_WIDTHS, REC_TEXT_COLS)
    _add_named_range(wb, ws_rec, RECEIPT_HEADERS, len(receipt_rows), "DatabaseReceipts")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_allocation_report(allocation_rows: list) -> io.BytesIO:
    """קובץ Excel לשליחה ללקוחה — השלמת ת.ז ומספרי הקצאה"""
    wb = Workbook()
    ws = wb.active
    ws.title = "דוח הקצאה"

    # כותרת
    ws.append(ALLOCATION_HEADERS)
    for i, cell in enumerate(ws[1], 1):
        cell.fill      = PatternFill("solid", fgColor="1E4D2B")
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(1)].width = 14
    ws.column_dimensions[get_column_letter(2)].width = 13
    ws.column_dimensions[get_column_letter(3)].width = 28
    ws.column_dimensions[get_column_letter(4)].width = 16
    ws.column_dimensions[get_column_letter(5)].width = 18
    ws.column_dimensions[get_column_letter(6)].width = 18
    ws.row_dimensions[1].height = 22

    # נתונים
    for row in allocation_rows:
        ws.append(row)

    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, len(ALLOCATION_HEADERS) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if c_idx == 2:                         # תאריך — TEXT
                if hasattr(cell.value, 'strftime'):
                    cell.value = f"{cell.value.day:02d}/{cell.value.month:02d}/{cell.value.year}"
                cell.number_format = "@"
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 5:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="left")
            elif c_idx == 4 and not cell.value:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")  # צהוב = ריק למילוי
            elif c_idx == 6:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")  # צהוב = ריק למילוי

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
