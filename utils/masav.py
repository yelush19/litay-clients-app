import io
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, quote_sheetname

GREEN = '4CAF50'

def format_date(dt):
    return f"{dt.day:02d}/{dt.month:02d}/{dt.year}"

def parse_amount(s):
    import re
    c = re.sub(r"[\$,\s]","",str(s or ""))
    c = re.sub(r"[()]","",c)
    try: return abs(float(c))
    except: return 0.0

BLUE  = '1565C0'

# ===== MASAV PARSER + BUILDER =====
def read_vendor_index(index_json):
    return index_json if isinstance(index_json, dict) else {}

def parse_masav(file_bytes, vendor_lookup, bank_coa):
    rows,batches,errors,unmatched=[],{},[],[]
    try:
        wb=load_workbook(io.BytesIO(file_bytes),data_only=True)
        ws=wb.active
        header=[str(c.value or '') for c in list(ws.iter_rows(min_row=1,max_row=1))[0]]
        def col(name): return next((i for i,h in enumerate(header) if name in h),None)
        cv=col('שם ספק'); ca=col('סכום'); chp=col('ח.פ'); cd=col('תאריך')
        cpd=col('שולם'); cb=col('Batch'); ci=col('חשבונית')
        cbk=col('בנק'); cbr=col('מספר סניף'); cac=col('מספר חשבון')
        for ri,row in enumerate(ws.iter_rows(min_row=2,max_row=ws.max_row,values_only=True),2):
            vendor=str(row[cv] or '').strip()
            if not vendor or vendor=='סך הכל לתשלום': continue
            paid=row[cpd]
            if paid is not True and str(paid).upper() not in ('TRUE','1','כן'): continue
            raw_date=row[cd]
            dt=raw_date if isinstance(raw_date,datetime) else None
            if not dt:
                for fmt in ('%d/%m/%Y','%Y-%m-%d','%d.%m.%Y'):
                    try: dt=datetime.strptime(str(raw_date),fmt); break
                    except: pass
            if not dt: errors.append(f"שורה {ri}: תאריך לא תקין"); continue
            try: amount=float(str(row[ca] or '').replace(',','').replace('₪','').strip())
            except: errors.append(f"שורה {ri}: סכום לא תקין"); continue
            hp_raw=row[chp]
            hp_raw_str = str(int(hp_raw)) if isinstance(hp_raw,float) else str(hp_raw or '').strip()
            # נרמול: הסר אפסים מובילים
            hp = hp_raw_str.lstrip('0') or hp_raw_str
            invoice=str(row[ci] or '') if ci is not None else ''
            batch=str(row[cb] or 'ללא batch') if cb is not None else 'ללא batch'
            bank_d=f"{row[cbk] or ''}-{row[cbr] or ''}-{row[cac] or ''}" if cbk is not None else ''
            vendor_coa=vendor_lookup.get(hp,'')
            if not vendor_coa:
                unmatched.append({'vendor':vendor,'hp':hp,'amount':amount,
                                  'date':format_date(dt)})
            rec={'date':dt,'date_fmt':format_date(dt),'vendor':vendor,'amount':amount,
                 'invoice':invoice,'hp':hp,'vendor_coa':vendor_coa,'bank_coa':bank_coa,
                 'bank_detail':bank_d,'batch':batch}
            rows.append(rec)
            if batch not in batches: batches[batch]={'total':0,'count':0}
            batches[batch]['total']+=amount; batches[batch]['count']+=1
    except Exception as e:
        errors.append(f"שגיאה: {e}")
    rows.sort(key=lambda r:(r['batch'],r['date']))
    return rows,batches,errors,unmatched

def _clean_ref(val, max_len=9) -> str:
    """אסמכתא — ספרות בלבד, טקסט, מקסימום 9 תווים"""
    import re as _re
    digits = _re.sub(r'[^0-9]', '', str(val or ''))
    return digits[:max_len]


def build_masav_excel(rows, batches):
    wb = Workbook()
    ws = wb.active
    ws.title = 'MASAV'

    headers = ['תאריך','תיאור','חובה','זכות',
               'אסמכתא 1','אסמכתא 2',
               'ח"ן חובה','ח"ן זכות','Batch','פרטי בנק ספק']
    ws.append(headers)

    hf    = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="right")

    for col, w in zip("ABCDEFGHIJ", [12,35,14,14,12,12,12,12,10,26]):
        ws.column_dimensions[col].width = w

    # ===== שורות נתונים בלבד — Named Range יכסה רק אותן =====
    data_start_row = 2

    for t in rows:
        # תאריך — טקסט DD/MM/YYYY
        date_val = str(t['date_fmt'])

        # אסמכתאות — ספרות בלבד, מקס 9 תווים, TEXT
        ref1 = _clean_ref(t['invoice'], 9)
        ref2 = _clean_ref(t['hp'], 9)

        ws.append([date_val, t['vendor'], t['amount'], t['amount'],
                   ref1, ref2,
                   str(t['vendor_coa'] or ''),
                   str(t['bank_coa'] or ''),
                   str(t['batch'] or ''),
                   str(t['bank_detail'] or '')])

        r = ws.max_row

        # תאריך — TEXT
        ws.cell(r, 1).number_format = '@'

        # סכומים
        for c in (3, 4):
            ws.cell(r, c).number_format = '#,##0.00'

        # אסמכתאות + ח"ן + Batch — TEXT
        for c in (5, 6, 7, 8, 9):
            cell = ws.cell(r, c)
            cell.number_format = '@'

        # ח"ן חובה חסר — צהוב
        if not t['vendor_coa']:
            ws.cell(r, 7).fill = PatternFill(
                start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

    # Named Range — רק שורות נתונים (ללא סיכומים)
    last_data_row = ws.max_row
    if rows:
        ref = f"{quote_sheetname('MASAV')}!$A${data_start_row}:$J${last_data_row}"
        wb.defined_names.add(DefinedName('MASAV', attr_text=ref))

    # ===== סיכומים — מתחת ל-Named Range =====
    current_batch = None
    batch_start   = data_start_row

    # חישוב סיכומי batch מהנתונים
    batch_totals = {}
    row_num = data_start_row
    for t in rows:
        b = str(t['batch'])
        if b not in batch_totals:
            batch_totals[b] = {'start': row_num, 'end': row_num}
        else:
            batch_totals[b]['end'] = row_num
        row_num += 1

    # שורות סיכום batch
    for b, info in batch_totals.items():
        s, e = info['start'], info['end']
        ws.append(['', f'סך הכל batch {b}',
                   f'=SUM(C{s}:C{e})', f'=SUM(D{s}:D{e})',
                   '', '', '', '', '', ''])
        r = ws.max_row
        for c in (3, 4): ws.cell(r, c).number_format = '#,##0.00'
        sf = Font(bold=True)
        sfill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        for c in range(1, 11):
            ws.cell(r, c).font  = sf
            ws.cell(r, c).fill  = sfill

    # שורת סיכום כולל
    total = sum(t['amount'] for t in rows)
    tr = ws.max_row + 1
    ws.append(['', 'סה״כ כולל', total, total, '', '', '', '', '', ''])
    for c in (3, 4): ws.cell(tr, c).number_format = '#,##0.00'
    tf    = Font(bold=True, color='FFFFFF')
    tfill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type='solid')
    for c in range(1, 11):
        ws.cell(tr, c).font = tf
        ws.cell(tr, c).fill = tfill

    no_coa = sum(1 for r in rows if not r['vendor_coa'])
    return wb, {
        'total_rows':   len(rows),
        'total_amount': total,
        'no_coa':       no_coa,
        'coa_covered':  len(rows) - no_coa,
    }

# ===== FUZZY MATCH VENDORS =====
from difflib import SequenceMatcher

def fuzzy_match_vendor(masav_name: str, vendor_index: dict) -> tuple:
    """
    מנסה למצוא שם דומה מאינדקס הספקים לפי שם (לא ח.פ.)
    מחזיר (שם_מוצע, ח"ן, ratio) או (None, None, 0)
    vendor_index: {ח.פ. → {name: שם, account: ח"ן}}
    """
    name_lower = masav_name.lower().strip()
    name_words = name_lower.split()
    
    best_name    = None
    best_account = None
    best_hp      = None
    best_ratio   = 0

    for hp, data in vendor_index.items():
        if not isinstance(data, dict):
            continue
        candidate = str(data.get("name", "")).lower().strip()
        if not candidate:
            continue

        # כל המילים של השם הקצר מופיעות בשם הארוך
        if len(name_words) >= 2 and all(w in candidate for w in name_words):
            ratio = 0.9
        elif len(name_words) >= 2 and all(w in name_lower for w in candidate.split()):
            ratio = 0.85
        else:
            ratio = SequenceMatcher(None, name_lower, candidate).ratio()

        if ratio >= 0.7 and ratio > best_ratio:
            best_ratio   = ratio
            best_name    = data.get("name")
            best_account = data.get("account")
            best_hp      = hp

    return best_name, best_account, best_hp, best_ratio


def build_vendor_index_with_names(vendor_index_raw: dict, coa_lookup: dict) -> dict:
    """
    ממיר vendor_index פשוט {ח.פ. → ח"ן} ל-{ח.פ. → {name, account}}
    coa_lookup: מ-Supabase, כולל שמות אם יש
    """
    result = {}
    for hp, value in vendor_index_raw.items():
        if isinstance(value, dict):
            result[hp] = value
        else:
            # פורמט ישן {ח.פ. → ח"ן}
            result[hp] = {"account": str(value), "name": ""}
    return result


def read_vendor_index_xlsx(file_bytes: bytes) -> tuple:
    """
    קורא אינדקס ספקים מחשבשבת (XLSX).
    מחזיר ({ח.פ. → ח"ן}, errors)
    קריאה דרך zipfile/XML כדי להתמודד עם borderID
    """
    lookup = {}
    errors = []

    def norm_hp(hp: str) -> str:
        """נרמול ח.פ. — הסר אפסים מובילים"""
        return hp.lstrip('0') or hp

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            shared = []
            if 'xl/sharedStrings.xml' in z.namelist():
                ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
                for si in ET.parse(z.open('xl/sharedStrings.xml')).getroot():
                    parts = si.findall(f'.//{ns}t')
                    shared.append(''.join(p.text or '' for p in parts))

            sheet_name = next(
                (n for n in z.namelist() if 'worksheets/sheet' in n and n.endswith('.xml')),
                None
            )
            if not sheet_name:
                errors.append("לא נמצא גיליון בקובץ")
                return lookup, errors

            sheet = ET.parse(z.open(sheet_name))
            ns = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'

            for row in sheet.getroot().findall(f'.//{ns}row'):
                cells = {}
                for c in row.findall(f'{ns}c'):
                    ref = c.get('r', '')
                    col = ''.join(filter(str.isalpha, ref))
                    t   = c.get('t', '')
                    v   = c.find(f'{ns}v')
                    val = v.text if v is not None else ''
                    if t == 's' and val:
                        val = shared[int(val)]
                    cells[col] = val

                # E=קוד מיון(300), F=מפתח חשבון (ספרות בלבד עד 7 תווים), I=ח.פ.
                if cells.get('E') == '300' and cells.get('F') and cells.get('F') not in ('', '300', 'מפתח חשבון'):
                    hp      = cells.get('I', '').strip()
                    account = cells.get('F', '').strip()
                    # סנן מזהים פנימיים של חשבשבת (לא ספרות או יותר מ-7 תווים)
                    if not account.isdigit() or len(account) > 7:
                        continue
                    if hp and hp not in ('0', '', '999999999') and account:
                        lookup[norm_hp(hp)] = account

    except Exception as e:
        errors.append(f"שגיאה בקריאת אינדקס: {e}")

    return lookup, errors
