import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, quote_sheetname

GREEN = "4CAF50"
BLUE  = "1565C0"

# ===== HELPERS =====
def clean_description(bai_desc, detail):
    c = re.sub(r"^20\d{11,}\s*","",detail or "")
    c = re.sub(r"\d{10,}","",c)
    c = re.sub(r"\b\d{6}\b","",c)
    c = re.sub(r"ST-?[A-Z0-9]{10,}","",c)
    c = re.sub(r"\d{2}/\d{2}/\d{2}","",c)
    c = re.sub(r"\s+"," ",c).strip()[:40]
    return c if len(c)>=3 else bai_desc

def parse_amount(s):
    c = re.sub(r"[\$,\s]","",str(s or ""))
    c = re.sub(r"[()]","",c)
    try: return abs(float(c))
    except: return 0.0

def parse_net_amount(s):
    c = re.sub(r"[\$,\s]","",str(s or ""))
    neg = "(" in c
    c = re.sub(r"[()]","",c)
    try: v = float(c)
    except: v = 0.0
    return -v if neg else v

def format_date(dt):
    return f"{dt.day:02d}/{dt.month:02d}/{dt.year}"

def parse_valley_date(s):
    parts = str(s or "").split("/")
    if len(parts)!=3: return None
    try: return datetime(int(parts[2]),int(parts[0]),int(parts[1]))
    except: return None

def parse_payem_date(s):
    parts = str(s or "").split("-")
    if len(parts)!=3: return None
    try: return datetime(int(parts[0]),int(parts[1]),int(parts[2]))
    except: return None

def detect_file_type(rows):
    if not rows or len(rows)<2: return "unknown"
    r0=[str(c or "") for c in rows[0]]
    r1=[str(c or "") for c in rows[1]]
    if any("Valley" in c for c in r0): return "valley"
    if any("PayEm" in c for c in r0): return "payem"
    if r1 and r1[0].startswith("Bank"): return "valley"
    if len(r1)>=3 and r1[0]=="Date" and r1[1]=="Time" and r1[2]=="Status": return "payem"
    return "unknown"

def workbook_to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def _format_text_col(ws,col,min_row,max_row):
    for row in ws.iter_rows(min_row=min_row,max_row=max_row,min_col=col,max_col=col):
        for cell in row:
            if cell.value is not None:
                cell.value=str(cell.value)
                cell.number_format='@'

def _add_named_range(wb,name,ws_title,min_col,min_row,max_col,max_row):
    ref=f"{quote_sheetname(ws_title)}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}"
    wb.defined_names.add(DefinedName(name,attr_text=ref))

GREEN = "4CAF50"
BLUE  = "1565C0"

# ===== VALLEY PARSER + BUILDER =====
def parse_valley(rows, coa_lookup, cat_coa):
    results=[]
    for i in range(2,len(rows)):
        row=rows[i]
        if not row or len(row)<12: continue
        if str(row[0] or "").startswith("Bank"): continue
        date_str=str(row[5] or "")
        type_ind=str(row[6] or "").upper()
        amount_str=str(row[7] or "")
        import re as _re3
        cust_ref = _re3.sub(r'[^0-9]', '', str(row[8] or ''))[:9]
        bai_desc=str(row[10] or "").upper()
        detail=str(row[11] or "")
        category=str(row[20] or "") if len(row)>20 else ""
        is_valid=("ACH" in bai_desc or "WIRE" in bai_desc or "DEPOSIT" in bai_desc
                  or "FEE" in bai_desc or type_ind in ("DEBIT","CREDIT"))
        if not is_valid or "/" not in date_str or not amount_str: continue
        dt=parse_valley_date(date_str)
        if not dt: continue
        amount=parse_amount(amount_str)
        desc=clean_description(bai_desc,detail)
        r=str(cust_ref).strip()
        key=f"{date_str}|{amount}|{r}" if r and r!="0" else f"{date_str}|{amount}|{bai_desc}|{detail[:20]}"
        coa_key=f"{desc}|{category}"
        coa_m=coa_lookup.get(coa_key)
        cat_m=cat_coa.get(category)
        results.append({"key":key,"date":dt,"date_formatted":format_date(dt),
            "description":desc,"amount":amount,"category":category,
            "coa_debit":coa_m[0] if coa_m else (cat_m[0] if cat_m else ""),
            "coa_credit":coa_m[1] if coa_m else (cat_m[1] if cat_m else ""),
            "type_ind":type_ind})
    results.sort(key=lambda t:t["date"])
    return results

def build_valley_excel(data):
    wb=Workbook()
    ws=wb.active
    ws.title="VALLEYTRANS"
    headers=["תאריך","תיאור","חובה","זכות","קטגוריה",'ח"ן חובה','ח"ן זכות']
    ws.append(headers)
    hf=PatternFill(start_color=GREEN,end_color=GREEN,fill_type="solid")
    hfont=Font(bold=True,color="FFFFFF")
    for cell in ws[1]:
        cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal="right")
    for t in data:
        ws.append([t["date_formatted"],t["description"],t["amount"],t["amount"],
                   t["category"],t["coa_debit"],t["coa_credit"]])
    for row in ws.iter_rows(min_row=2,min_col=3,max_col=4):
        for cell in row: cell.number_format='#,##0.00'
    _format_text_col(ws,6,2,ws.max_row)
    _format_text_col(ws,7,2,ws.max_row)
    for col,w in zip("ABCDEFG",[12,42,14,14,28,12,12]):
        ws.column_dimensions[col].width=w
    if data: _add_named_range(wb,"VALLEYTRANS","VALLEYTRANS",1,1,7,ws.max_row)
    # דף בנק
    ws2=wb.create_sheet(title="דף בנק")
    ws2.append(["תאריך","תאריך ערך","תיאור","חובה מטח","זכות מטח","יתרה"])
    hf2=PatternFill(start_color=BLUE,end_color=BLUE,fill_type="solid")
    for cell in ws2[1]: cell.fill=hf2; cell.font=hfont; cell.alignment=Alignment(horizontal="right")
    balance=0.0
    for t in data:
        is_debit=t.get("type_ind","DEBIT")=="DEBIT"
        dv=t["amount"] if is_debit else None
        cv=t["amount"] if not is_debit else None
        balance += (-t["amount"] if is_debit else t["amount"])
        ws2.append([t["date_formatted"],t["date_formatted"],t["description"],dv,cv,balance])
    for row in ws2.iter_rows(min_row=2,min_col=4,max_col=6):
        for cell in row:
            if cell.value is not None: cell.number_format='#,##0.00'
    for col,w in zip("ABCDEF",[12,12,42,14,14,16]):
        ws2.column_dimensions[col].width=w
    if data: _add_named_range(wb,"VALLEY_BANK","דף בנק",1,1,6,ws2.max_row)
    total_debits=sum(t["amount"] for t in data if t.get("type_ind")=="DEBIT")
    total_credits=sum(t["amount"] for t in data if t.get("type_ind")!="DEBIT")
    total=sum(t["amount"] for t in data)
    no_coa=sum(1 for t in data if not t["coa_debit"])
    return wb,{"amount_ok":abs(total-(total_debits+total_credits))<0.01,
               "total_amount":total,"total_debits":total_debits,"total_credits":total_credits,
               "expected_balance":total_credits-total_debits,
               "no_coa":no_coa,"coa_covered":len(data)-no_coa,"total_count":len(data)}

# ===== PAYEM PARSER + BUILDER =====
def parse_payem(rows,coa_lookup,ltd_coa):
    results=[]
    for i in range(2,len(rows)):
        row=rows[i]
        if not row or len(row)<42: continue
        if str(row[2] or "").upper()!="CLEARED": continue
        dt=parse_payem_date(str(row[0] or ""))
        if not dt: continue
        merchant=str(row[28] or "").strip()
        txn_id=str(row[41] or "").strip()
        import re as _re2
        _card_raw = str(row[35] or '').strip()
        card4 = _re2.sub(r'[^0-9]', '', _card_raw)[:9]
        subsidiary=str(row[9] or "").strip()
        net_str=str(row[53] or "") if len(row)>53 else str(row[5] or "")
        net=parse_net_amount(net_str)
        abs_amt=abs(net)
        is_cr=net>0
        coa_key=f"{merchant}|{subsidiary}"
        coa_m=coa_lookup.get(coa_key)
        ltd_m=ltd_coa.get(merchant)
        # אסמכתא 1: ספרות בלבד, מקס 9 תווים, TEXT
        import re as _re
        ref1 = _re.sub(r'[^0-9]', '', txn_id)[:9]
        results.append({"key":txn_id,"date":dt,"date_formatted":format_date(dt),
            "description":merchant,"net_amount":net,"abs_amount":abs_amt,
            "credit_amount":abs_amt if is_cr else None,
            "debit_amount":abs_amt if not is_cr else None,
            "ref1":ref1,"ref2":card4,
            "coa_credit":coa_m[0] if coa_m else "",
            "coa_debit":coa_m[1] if coa_m else "",
            "ltd_coa_credit":ltd_m[0] if ltd_m else "",
            "ltd_coa_debit":ltd_m[1] if ltd_m else "",
            "subsidiary":subsidiary})
    results.sort(key=lambda t:t["date"])
    return results

def _add_payem_sheet(wb,data,sheet_name,mode):
    ws=wb.create_sheet(title=sheet_name)
    if mode=="all":
        headers=["תיאור","Transaction amount","זכות מטח","חובה מטח","אסמכתא 1","אסמכתא 2",'ח"ן זכות','ח"ן חובה',"תאריך","שיוך"]
    else:
        headers=["תיאור","זכות מטח","חובה מטח","אסמכתא 1","אסמכתא 2",'ח"ן זכות','ח"ן חובה',"תאריך","שיוך"]
    ws.append(headers)
    hf=PatternFill(start_color=GREEN,end_color=GREEN,fill_type="solid")
    hfont=Font(bold=True,color="FFFFFF")
    for cell in ws[1]: cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal="right")
    for t in data:
        is_cr=t["net_amount"]>0
        if mode=="ltd": bcr=t["ltd_coa_credit"] or ""; bdr=t["ltd_coa_debit"] or ""
        elif mode=="negdit": bcr="502010"; bdr="300001"
        elif mode=="inc": bcr=t["coa_credit"]; bdr=t["coa_debit"]
        else: bcr=t["coa_credit"]; bdr=t["coa_debit"]
        if mode!="all" and is_cr: coa_cr=bdr; coa_dr=bcr
        else: coa_cr=bcr; coa_dr=bdr
        if mode=="all":
            ws.append([t["description"],t["net_amount"],t["credit_amount"],t["debit_amount"],
                       t["ref1"],t["ref2"],coa_cr,coa_dr,t["date"],t["subsidiary"]])
        else:
            ws.append([t["description"],t["abs_amount"],t["abs_amount"],
                       t["ref1"],t["ref2"],coa_cr,coa_dr,t["date"],t["subsidiary"]])
    if mode=="all":
        nc=(2,4); r1c=5; r2c=6; crcol=7; drcol=8; dcol=9; tcols=10
        cw={"A":30,"B":16,"C":14,"D":14,"E":14,"F":10,"G":12,"H":12,"I":12,"J":22}
    else:
        nc=(2,3); r1c=4; r2c=5; crcol=6; drcol=7; dcol=8; tcols=9
        cw={"A":30,"B":14,"C":14,"D":14,"E":10,"F":12,"G":12,"H":12,"I":22}
    for row in ws.iter_rows(min_row=2,min_col=nc[0],max_col=nc[1]):
        for cell in row:
            if cell.value is not None: cell.number_format='#,##0.00'
    for row in ws.iter_rows(min_row=2,min_col=dcol,max_col=dcol):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.value = cell.value.strftime('%d/%m/%Y')
                cell.number_format = '@'
            elif cell.value is not None:
                cell.number_format = '@'
    for c in (r1c,r2c,crcol,drcol): _format_text_col(ws,c,2,ws.max_row)
    for col,w in cw.items(): ws.column_dimensions[col].width=w
    if data:
        nm={"PAYEMDATA":"PAYEMDATA","פקודה_INC":"PEKUDA_INC","רישום_LTD":"RISHUM_LTD","נגדית_לINC_של_LTD":"NEGDIT_LTD"}
        sn=sheet_name.replace(" ","_").replace('"','')
        rn=nm.get(sn,sn)
        try: _add_named_range(wb,rn,sheet_name,1,1,tcols,ws.max_row)
        except: pass

def build_payem_excel(data):
    wb=Workbook(); wb.remove(wb.active)
    inc=[t for t in data if "Inc" in t["subsidiary"]]
    ltd=[t for t in data if "LTD" in t["subsidiary"]]
    _add_payem_sheet(wb,data,"PAYEMDATA","all")
    _add_payem_sheet(wb,inc,"פקודה INC","inc")
    _add_payem_sheet(wb,ltd,"רישום LTD","ltd")
    _add_payem_sheet(wb,ltd,"נגדית לINC של LTD","negdit")
    dd=sum(t["abs_amount"] for t in data if t["net_amount"]<=0)
    dc=sum(t["abs_amount"] for t in data if t["net_amount"]>0)
    id_=sum(t["abs_amount"] for t in inc if t["net_amount"]<=0)
    ic=sum(t["abs_amount"] for t in inc if t["net_amount"]>0)
    ld=sum(t["abs_amount"] for t in ltd if t["net_amount"]<=0)
    lc=sum(t["abs_amount"] for t in ltd if t["net_amount"]>0)
    return wb,{"count_ok":len(inc)+len(ltd)==len(data),
               "count_data":len(data),"count_inc":len(inc),"count_ltd":len(ltd),
               "debit_ok":abs(dd-(id_+ld))<0.01,"data_debit":dd,"inc_debit":id_,"ltd_debit":ld,
               "credit_ok":abs(dc-(ic+lc))<0.01,"data_credit":dc,"inc_credit":ic,"ltd_credit":lc}